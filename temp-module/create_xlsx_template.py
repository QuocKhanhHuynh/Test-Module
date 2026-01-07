
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

wb = Workbook()
ws = wb.active
ws.title = "OCR Results"

# Multi-level headers
header_groups = [
    ("Tổng thời gian", ["task 1", "task 2", "task 3"]),
    ("Nhận diện được QR không", ["task 1", "task 2", "task 3"]),
    ("Giá trị QR", ["task 1", "% chính xác task 1", "task 2", "% chính xác task 2", "task 3", "% chính xác task 3", "chuẩn"]),
    ("Giá trị tổng số lượng", ["task 1", "% chính xác task 1", "task 2", "% chính xác task 2", "task 3", "% chính xác task 3", "chuẩn"]),
    ("Giá trị kiểu áo", ["task 1", "% chính xác task 1", "task 2", "% chính xác task 2", "task 3", "% chính xác task 3", "chuẩn"]),
    ("Giá trị size áo", ["task 1", "% chính xác task 1", "task 2", "% chính xác task 2", "task 3", "% chính xác task 3", "chuẩn"]),
    ("Giá trị màu áo", ["task 1", "% chính xác task 1", "task 2", "% chính xác task 2", "task 3", "% chính xác task 3", "chuẩn"])
]

# Tạo dòng header 1 (group)
group_headers = []
for group, cols in header_groups:
    group_headers.extend([group] * len(cols))
ws.append(group_headers)

# Tạo dòng header 2 (sub)
sub_headers = []
for group, cols in header_groups:
    sub_headers.extend(cols)
ws.append(sub_headers)

# Merge các ô cho group header
col_idx = 1
for group, cols in header_groups:
    if len(cols) > 1:
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+len(cols)-1)
        cell = ws.cell(row=1, column=col_idx)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    col_idx += len(cols)

wb.save("OCR_Result_Template.xlsx")
print("Đã tạo file OCR_Result_Template.xlsx với multi-level header và cột % chính xác.")
